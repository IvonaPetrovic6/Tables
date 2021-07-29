import openpyxl
from openpyxl.chart import PieChart3D, Reference

def create_charts(file):                                                                                     # method create_charts is defined
    table = openpyxl.load_workbook(file)                                                                     # it takes the name of xlsx file as an argument
    list = table["List1"]                                                                                    # program works with xlsx files which contain table
    sum = [0,0,0,0,0,0]                                                                                      # when calling this function, 3D pi charts are created
                                                                                                             # for every column of the table
    for column in range(2,list.max_column + 1):
        for row in range(2, list.max_row + 1):                                                               # this nested loop counts the sum of every column
            cell = list.cell(row,column)                                                                     # and stores it in the list sum[]
            sum[column - 2] += cell.value

    for column in range(2,list.max_column + 1):
        for row in range(2, list.max_row + 1):
            cell = list.cell(row,column)                                                                     # this nested loop creates percentage (instead
            percentage = cell.value / sum[column - 2]                                                        # of numbers) for every cell in the table
            cell.value = percentage * 100                                                                    # and rounds each number to four decimals
            four_decimals = "{:.4f}".format(cell.value)
            cell.value = f"{four_decimals}%"

    table.save(file)                                                                                         # new table (with percentage) is now saved in xlsx file

    for column in range(2,list.max_column + 1):
        values = Reference(list, min_row = 2, max_row = list.max_row, min_col = column, max_col = column)
        chart = PieChart3D()                                                                                 # this loop creates a chart for every column in the table
        chart.add_data(values)                                                                               # so every chart shows the percentage of data that
        cell = list.cell(1,column)                                                                           # imported table contains
        chart.title = cell.value                                                                             # loop also adds the title to each chart which is equal to
        list.add_chart(chart, f"a{list.max_row + 3 + 16 * (column - 2)}")                                    # the title of the column in table

    table.save(file)                                                                                         # charts are also saved in xlsx file with the table


create_charts("Population_in_millions.xlsx")                                                                 # calling the method create_charts and passing
                                                                                                             # Population_in_millions.xlsx file to that method