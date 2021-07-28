import openpyxl
from openpyxl.chart import PieChart3D, Reference

table = openpyxl.load_workbook("Population_per_percentage.xlsx")                                         # we import this file because it contains data of
list = table["List1"]                                                                                    # which we will form charts

for column in range(2,8):
    values = Reference(list, min_row = 2, max_row = list.max_row, min_col = column, max_col = column)
    chart = PieChart3D()                                                                                 # this loop creates a chart for every column in the table
    chart.add_data(values)                                                                               # so every chart shows the percentage of population who
    cell = list.cell(1,column)                                                                           # belongs to certain religion on each continent
    chart.title = cell.value                                                                             # loop also adds the title to each chart which is equal to
    list.add_chart(chart, f"a{list.max_row + 3 + 16 * (column - 2)}")                                    # the title of the column in table (name of a continent)

table.save("Charts.xlsx")                                                                                # table with charts is saved in Charts.xlsx