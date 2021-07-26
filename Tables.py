import openpyxl

table = openpyxl.load_workbook("Population_in_millions.xlsx")
list = table["List1"]
sum = [0,0,0,0,0,0]

for column in range(2,list.max_column + 1):
    for row in range(2, list.max_row + 1):            # this nested loop counts the sum of every column
        cell = list.cell(row,column)                  # and stores it in the list sum[]
        sum[column - 2] += cell.value

for column in range(2,list.max_column + 1):
    for row in range(2, list.max_row + 1):
        cell = list.cell(row,column)                  # this nested loop creates percentage (instead of
        percentage = cell.value / sum[column - 2]     # population in millions) for every cell in the table
        cell.value = percentage * 100                 # and rounds each number to four decimals
        four_decimals = "{:.4f}".format(cell.value)
        cell.value = f"{four_decimals}%"

table.save("Population_per_percentage.xlsx")          # new table (with percentage) is stored in another xlsx file


